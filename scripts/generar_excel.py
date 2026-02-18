#!/usr/bin/env python3
"""
Generador de Excel profesional para compartir con la familia de Paulina.
Lee de los JSONs y genera un Excel formateado y bonito.
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

# Rutas
BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / "output"

# Estilos
HEADER_FILL = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TOTAL_FILL = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
CURRENCY_FORMAT = '€#,##0'
PERCENT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def aplicar_estilo_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

def aplicar_estilo_subheader(cell):
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

def aplicar_estilo_total(cell):
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.border = THIN_BORDER

def aplicar_borde(cell):
    cell.border = THIN_BORDER

def cargar_datos():
    with open(OUTPUT_DIR / "datos_paulina.json", "r", encoding="utf-8") as f:
        datos = json.load(f)
    with open(OUTPUT_DIR / "escenarios_paulina.json", "r", encoding="utf-8") as f:
        escenarios = json.load(f)
    return datos, escenarios

def crear_hoja_resumen(wb, datos, escenarios):
    """Crea la hoja de resumen ejecutivo"""
    ws = wb.active
    ws.title = "Resumen Ejecutivo"

    # Titulo
    ws.merge_cells('A1:F1')
    ws['A1'] = "PRESUPUESTO DE PERMANENCIA - PAULINA EN MADRID"
    ws['A1'].font = Font(bold=True, size=16, color="1a365d")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"IE University | {datos['perfil']['ano_inicio']}-{datos['perfil']['ano_inicio'] + datos['perfil']['duracion_anos']}"
    ws['A2'].font = Font(size=12, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')

    # Informacion del perfil
    row = 4
    ws[f'A{row}'] = "PERFIL"
    aplicar_estilo_header(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:B{row}')

    perfil_data = [
        ("Nombre", datos['perfil']['nombre']),
        ("Edad", f"{datos['perfil']['edad']} anos"),
        ("Universidad", datos['perfil']['universidad']),
        ("Programa", datos['perfil']['programa']),
        ("Duracion", f"{datos['perfil']['duracion_anos']} anos"),
        ("Ciudad", datos['perfil']['ciudad']),
        ("Pais origen", datos['perfil']['pais_origen'])
    ]

    for label, value in perfil_data:
        row += 1
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        aplicar_borde(ws[f'A{row}'])
        aplicar_borde(ws[f'B{row}'])

    # Comparativa de escenarios
    row += 2
    ws[f'A{row}'] = "COMPARATIVA DE ESCENARIOS (4 ANOS)"
    aplicar_estilo_header(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:E{row}')

    row += 1
    headers = ["Escenario", "Total 4 Anos (EUR)", "Total (USD)", "Total (COP)", "Mensual Prom."]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        aplicar_estilo_subheader(cell)

    for key in ["austero", "moderado", "comodo"]:
        row += 1
        esc = escenarios["escenarios"][key]
        ws.cell(row=row, column=1, value=key.upper())
        ws.cell(row=row, column=2, value=esc["totales"]["total_4_anos_eur"])
        ws.cell(row=row, column=3, value=esc["totales"]["total_4_anos_usd"])
        ws.cell(row=row, column=4, value=esc["totales"]["total_4_anos_cop"])
        ws.cell(row=row, column=5, value=esc["totales"]["promedio_mensual"])

        for col in range(1, 6):
            aplicar_borde(ws.cell(row=row, column=col))
            if col >= 2:
                ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT

    # Supuestos
    row += 2
    ws[f'A{row}'] = "SUPUESTOS"
    aplicar_estilo_header(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:B{row}')

    supuestos_data = [
        ("Matricula base anual", f"€{datos['costos_base']['matricula']['anual_base']:,}"),
        ("Descuento beca disponible", f"{datos['supuestos']['descuento_matricula_disponible']*100:.0f}%"),
        ("Inflacion estimada Espana", f"{datos['supuestos']['inflacion_espana']*100:.0f}%"),
        ("Tasa EUR/USD", datos['supuestos']['tasas_cambio']['EUR_USD']),
        ("Tasa EUR/COP", f"{datos['supuestos']['tasas_cambio']['EUR_COP']:,}")
    ]

    for label, value in supuestos_data:
        row += 1
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        aplicar_borde(ws[f'A{row}'])
        aplicar_borde(ws[f'B{row}'])

    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18

    return ws

def crear_hoja_escenario(wb, nombre, escenario, datos):
    """Crea una hoja detallada para cada escenario"""
    ws = wb.create_sheet(title=nombre.capitalize())

    # Titulo
    ws.merge_cells('A1:E1')
    ws['A1'] = f"ESCENARIO {nombre.upper()}"
    ws['A1'].font = Font(bold=True, size=14, color="1a365d")

    ws.merge_cells('A2:E2')
    ws['A2'] = escenario['descripcion']
    ws['A2'].font = Font(italic=True, color="666666")

    # Resumen ano 1
    row = 4
    ws[f'A{row}'] = "RESUMEN ANO 1"
    aplicar_estilo_header(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:C{row}')

    resumen = escenario['resumen_ano_1']
    resumen_data = [
        ("Matricula base", resumen['matricula_base']),
        ("Matricula con descuento", resumen['matricula_con_descuento']),
        ("Ahorro por descuento", resumen['ahorro_por_descuento']),
        ("Gastos mensuales", resumen['gastos_mensuales']),
        ("Gastos vida anual", resumen['gastos_vida_anual']),
        ("Vuelos anual", resumen['vuelos_anual']),
        ("Emergencias anual", resumen['emergencias_anual']),
        ("TOTAL ANUAL", resumen['total_anual']),
        ("Total mensual promedio", resumen['total_mensual_promedio'])
    ]

    for label, value in resumen_data:
        row += 1
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        aplicar_borde(ws[f'A{row}'])
        aplicar_borde(ws[f'B{row}'])
        if label == "TOTAL ANUAL":
            aplicar_estilo_total(ws[f'A{row}'])
            aplicar_estilo_total(ws[f'B{row}'])

    # Desglose mensual
    row += 2
    ws[f'A{row}'] = "DESGLOSE MENSUAL"
    aplicar_estilo_header(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    headers = ["Categoria", "Monto (EUR)", "Incluido"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        aplicar_estilo_subheader(cell)

    for cat, info in escenario['desglose_mensual'].items():
        row += 1
        ws.cell(row=row, column=1, value=info['descripcion'])
        ws.cell(row=row, column=2, value=info['valor'])
        ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=3, value="Si" if info['incluido'] else "No")

        for col in range(1, 4):
            aplicar_borde(ws.cell(row=row, column=col))

    # Proyeccion 4 anos
    row += 2
    ws[f'A{row}'] = "PROYECCION 4 ANOS"
    aplicar_estilo_header(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:E{row}')

    row += 1
    headers = ["Ano", "Matricula", "Gastos Vida", "Emergencias", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        aplicar_estilo_subheader(cell)

    start_data_row = row + 1
    for proy in escenario['proyeccion_anual']:
        row += 1
        ws.cell(row=row, column=1, value=proy['ano'])
        ws.cell(row=row, column=2, value=proy['matricula'])
        ws.cell(row=row, column=3, value=proy['gastos_vida'])
        ws.cell(row=row, column=4, value=proy['emergencias'])
        ws.cell(row=row, column=5, value=proy['total_anual'])

        for col in range(1, 6):
            aplicar_borde(ws.cell(row=row, column=col))
            if col >= 2:
                ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT

    # Fila de totales
    row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=f"=SUM(B{start_data_row}:B{row-1})")
    ws.cell(row=row, column=3, value=f"=SUM(C{start_data_row}:C{row-1})")
    ws.cell(row=row, column=4, value=f"=SUM(D{start_data_row}:D{row-1})")
    ws.cell(row=row, column=5, value=f"=SUM(E{start_data_row}:E{row-1})")

    for col in range(1, 6):
        aplicar_estilo_total(ws.cell(row=row, column=col))
        if col >= 2:
            ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT

    # Totales finales
    row += 2
    totales = escenario['totales']
    ws[f'A{row}'] = "TOTALES EN DIFERENTES MONEDAS"
    aplicar_estilo_header(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:B{row}')

    totales_data = [
        ("Total 4 anos (EUR)", totales['total_4_anos_eur']),
        ("Total 4 anos (USD)", totales['total_4_anos_usd']),
        ("Total 4 anos (COP)", totales['total_4_anos_cop']),
        ("Promedio anual", totales['promedio_anual']),
        ("Promedio mensual", totales['promedio_mensual'])
    ]

    for label, value in totales_data:
        row += 1
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = CURRENCY_FORMAT if "COP" not in label else '#,##0'
        aplicar_borde(ws[f'A{row}'])
        aplicar_borde(ws[f'B{row}'])

    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18

    return ws

def crear_hoja_costos_base(wb, datos):
    """Crea hoja con referencia de costos base"""
    ws = wb.create_sheet(title="Referencia Costos")

    ws.merge_cells('A1:E1')
    ws['A1'] = "REFERENCIA DE COSTOS - GUIA IE MADRID"
    ws['A1'].font = Font(bold=True, size=14, color="1a365d")

    row = 3
    headers = ["Categoria", "Descripcion", "Minimo", "Medio", "Maximo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        aplicar_estilo_header(cell)

    for cat, info in datos['costos_base'].items():
        if cat == "matricula" or cat == "emergencias":
            continue

        row += 1
        ws.cell(row=row, column=1, value=cat.replace("_", " ").title())
        ws.cell(row=row, column=2, value=info.get('descripcion', ''))
        ws.cell(row=row, column=3, value=info.get('min', '-'))
        ws.cell(row=row, column=4, value=info.get('medio', '-'))
        ws.cell(row=row, column=5, value=info.get('max', '-'))

        for col in range(1, 6):
            aplicar_borde(ws.cell(row=row, column=col))
            if col >= 3:
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = CURRENCY_FORMAT

    # Notas
    row += 2
    ws[f'A{row}'] = "NOTAS:"
    ws[f'A{row}'].font = Font(bold=True)

    for nota in datos['notas']:
        row += 1
        ws[f'A{row}'] = f"• {nota}"
        ws.merge_cells(f'A{row}:E{row}')

    # Ajustar anchos
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    return ws

def main():
    print("=" * 60)
    print("GENERADOR DE EXCEL PROFESIONAL - PAULINA MADRID")
    print("=" * 60)

    # Cargar datos
    print("\n[1/4] Cargando datos de JSONs...")
    datos, escenarios = cargar_datos()

    # Crear workbook
    print("[2/4] Creando Excel...")
    wb = Workbook()

    # Crear hojas
    print("[3/4] Generando hojas...")
    crear_hoja_resumen(wb, datos, escenarios)

    for nombre in ["moderado", "austero", "comodo"]:
        crear_hoja_escenario(wb, nombre, escenarios["escenarios"][nombre], datos)

    crear_hoja_costos_base(wb, datos)

    # Guardar
    output_path = OUTPUT_DIR / "resumen_paulina.xlsx"
    print(f"[4/4] Guardando en {output_path}...")
    wb.save(output_path)

    print("\n" + "=" * 60)
    print("Excel generado exitosamente!")
    print(f"Archivo: {output_path}")
    print("=" * 60)

    # Resumen de hojas
    print("\nHojas creadas:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

if __name__ == "__main__":
    main()
